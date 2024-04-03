import requests
import os
import json
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from datetime import date
from datetime import datetime

token_file = 'token.txt'
input_file = 'input.xlsx'

"""Список для фльтра"""
name_filter_one = 'ФГБУ'
name_filter_two = 'ФГБНУ'
name_filter_three = 'ФГБОУ'
name_filter_four = 'ФГУП'
name_filter_five = 'ФКП'
time_start = datetime.now()


def check_file(name_file):
    """Проверка на наличие файлов input.xlsx и token.txt"""
    if os.path.isfile(name_file):
        check = True
    else:
        check = False
    return check


def show_data_now():
    """Создание даты на момент проверки"""
    current_date = date.today()
    return current_date


def get_token(name_file):
    """Проверка наличия токена в файле"""
    if (check_file(name_file)):
        with open(name_file, 'r') as file_token:
            token_read = file_token.read()
        file_token.close()
    if ('' != token_read):
        token = token_read
    else:
        token = None
    return token


def get_input_adr(name_file):
    """Получение списка адресов из столбца файла input.xlsx"""
    wb = load_workbook(name_file)
    sheet = wb.active
    input_adr = []
    for cell in sheet['A2': 'A' + str(sheet.max_row)]:
        for value in cell:
            if value.value:
                input_adr.append(str(value.value))
    return input_adr


def wrtie_info_in_file_xls_pack(results):
    workbook = load_workbook("outpack.xlsx")  # Загрузка книги
    sheet = workbook.active
    sheet["A1"] = "Дата сверки"
    sheet["B1"] = "Название"
    sheet["C1"] = "Адрес организации"
    sheet["D1"] = "Название организации"
    sheet["E1"] = "Сайт"
    sheet["F1"] = "Телефон "
    sheet["G1"] = "Время работы"
    sheet["H1"] = "ID организации"
    sheet.column_dimensions['A'].width = 13
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 16
    for column in range(1, 9):
        sheet.cell(row=1, column=column).font = Font(bold=True)  # Установка жирного шрифта для названий столбцов
    for column in range(1, 9):
        sheet.cell(row=1, column=column).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Применение цвета к столбцам

    index = 2
    for adr, result in results.items():
        result_name_info = result.get('properties').get('ResponseMetaData').get('SearchRequest').get('request')

        for item in result.get('features'):
            property = item.get('properties')  # Извлечение свойств
            company_metadata = property.get("CompanyMetaData")
            address = company_metadata.get('address')  # "Адрес организации"
            name = property.get('name')  # "Название организации"
            if company_metadata.get('url') is None:
                contact_email = ''
            else:
                contact_email = company_metadata.get('url')  # "Сайт"
            contact_phone = company_metadata.get('Phones')
            if contact_phone is not None:
                phone_numbers_string = ''
                for phone in contact_phone:
                    phone_number = phone.get('formatted')
                    phone_numbers_string += f"{phone_number}"
                phone_numbers_info = phone_numbers_string[:-2]  # "Телефон"
            else:
                phone_numbers_info = ''
            if company_metadata.get('Hours') is not None:
                contact_work_time = company_metadata.get('Hours').get('text')  # "Время работы"
            else:
                contact_work_time = ''
            id_yandex = company_metadata.get('id')  # "ID организации"

            sheet.cell(row=index, column=1, value=datetime.now().strftime('%Y-%m-%d'))  # "Дата сверки"
            sheet.cell(row=index, column=2, value=result_name_info)  # "Название"
            sheet.cell(row=index, column=3, value=address)  # "Адрес организации"
            sheet.cell(row=index, column=4, value=name)  # "Название организации"
            sheet.cell(row=index, column=5, value=contact_email)  # "Сайт"
            sheet.cell(row=index, column=6, value=phone_numbers_info)  # "Телефон"
            sheet.cell(row=index, column=7, value=contact_work_time)  # "Время работы"
            sheet.cell(row=index, column=8, value=id_yandex)  # "ID организации"
            index += 1

    workbook.save("outpack.xlsx")  # Сохранение книги


def get_info_api(token, list_adr):
    """Получения данных с API Поиска по организациям"""
    results_to_write = {}  # Словарь для хранения результатов
    for adr in list_adr:
        url = 'https://search-maps.yandex.ru/v1/?text={}&type=biz&results=50&lang=ru_RU&apikey={}'.format(adr, token)
        res = requests.get(url)
        contents = res.text
        result = json.loads(contents)
        result_name_info = result.get('properties').get('ResponseMetaData').get('SearchRequest').get('request')
        print(result_name_info)
        result['input_adr'] = adr
        results_to_write[adr] = result  # Сохраняем результаты в словарь
    wrtie_info_in_file_xls_pack(results_to_write)  # После завершения цикла записываем результаты в файл


if __name__ == '__main__':
    check_file(input_file)  # Проверим наличие исходного файла с адресами
    token = get_token(token_file)  # Проверим наличия токена в файле
    adr_list = get_input_adr(input_file)  # Получим список исходных адресов
    get_info_api(token, adr_list)  # Запустим API
    show_data_now()  # wrtie_info_in_file()

time_end = datetime.now()
print("Время на обработку:", time_end - time_start)
